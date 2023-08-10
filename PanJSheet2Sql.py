import numpy as np
import tqdm
import re
import datetime
import sys
from typing import List, Tuple, Dict, Union, Optional, Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles.colors import Color
# input_path = sys.argv[1] #'D:\\C_KheuyMyenDong\\Desktop\\Jyutdict\\泛粵字表\\本体\\泛粵字表 220609.xlsx'
input_path = "Z:\\Proj\\Jyutdict\\泛粵字表\\本体\\泛粵字表 230806_.xlsx"
time_str = datetime.datetime.now().strftime('_%Y%m%d')

regex_pure_alphabet = re.compile("^[a-z'0-9?/①-⑨_^*]+$")
def str_format_enter_for_12(x: str) -> str:
    if len(regex_pure_alphabet.findall(x))>0: return x
    x = x.replace("；", ";")
    x = x.replace("，", ",")
    x = x.replace("：", ":")
    x = x.replace(", ", ",")
    x = x.replace("; ", ";")
    x = x.replace(",\n", ",")
    x = x.replace(";\n", ";")
    x = x.replace(",", ", ")
    x = x.replace(";", "; ")
    x = x.replace("  ", " ")
    x = x.replace(" :", ": ")
    x = x.replace("\n", "; ")
    x = x.replace("  ", " ")
    return x.strip()
def str_format_enter_for_0(x: str) -> str:
    x = x.replace("：需要例句", "")
    x = x.replace("⚠意味不明", "{？}").strip()
    if x[-1]=="$": x = x[:-1]
    return x


HEADER_INFO_COL_LENGTH = []
HEADER_INFO_COL_NAME = ['繁', '綜']
HEADER_INFO_MARK = [0, 0]
HEADER_INFO_FULL_NAME = [['字', ''], ['綜合音', '']]
HEADER_INFO_COLOR = ['', '']
HEADER_INFO_NOTE = ['', '']


regex_block_email = re.compile("@[a-zA-Z0-9\\.]+?@[a-zA-Z0-9]+?\\.[a-zA-Z]+\\b")
regex_block_name = re.compile("\n\t-.+?(\n|$)")
regex_block_info = lambda x: regex_block_name.sub("\t-[Anony]\n", regex_block_email.sub("@[Anony]", x.strip()) if"@"in x else x.strip())
def format_locale_record(row: int) -> Tuple[List[str], List[int], bool]:
    sheet_row = main_sheet[row]
    assert isinstance(sheet_row, tuple)
    valid = True
    result: List[str] = [""] * (HEADER_INFO_COL_COUNT+1)
    col_length = []
    locate_written_count = 0
    locate_owned_count = 0
    col_count = len(sheet_row)
    for n, item in enumerate(sheet_row):
        value = str(item.value).strip() if item.value else ""
        if value != "":
            if HEADER_INFO_MARK[n]==1:
                locate_written_count += 1
            if '"' in value and "'" in value:
                value = value.replace('"', "'")
            if value!="_" and HEADER_INFO_MARK[n]==1:
                locate_owned_count += 1
            if HEADER_INFO_MARK[n]!=0: value = str_format_enter_for_12(value)
            elif HEADER_INFO_MARK[n]==0: value = str_format_enter_for_0(value) 
            #if "原始" in HEADER_INFO_COL_NAME[n] and len(value)>1: value = "*"+value
            result[n] = ("'%s'" if "'" not in value else '"%s"') % value
        else:
            result[n] = "''"
    if locate_written_count<3 and locate_owned_count<2:
        valid = False
    
    if not valid:
        return result, [], False
    
    result[col_count] = '"{' + str({
        HEADER_INFO_COL_NAME[n] if item.comment else "": 
        #item.comment.text if item.comment else "" for n,item in enumerate(main_sheet[row])
        regex_block_info(item.comment.text) if item.comment else "" for n,item in enumerate(sheet_row)
    })[9:-1].replace("\"", "\\\"") + '}"'
    
    col_length = [len(item)-2 for item in result]
    return result, col_length, True
    
from colorsys import rgb_to_hls, hls_to_rgb
RGBMAX = 0xff  # Corresponds to 255
HLSMAX = 240  # MS excel's tint function expects that HLS is base 240.
def rgb_to_hex(red, green, blue):
    """Converts (0,1) based RGB values to a hex string 'rrggbb'"""
    # if green is None:
    #     red, green, blue = red
    return ('%02x%02x%02x' % (int(round(red * RGBMAX)), int(round(green * RGBMAX)), int(round(blue * RGBMAX)))).upper()
def ms_hls_to_rgb(hue, lightness, saturation) -> Tuple[float, float, float]:
    """Converts HLSMAX based HLS values to rgb values in the range (0,1)"""
    # if lightness is None:
    #     hue, lightness, saturation = hue
    return hls_to_rgb(hue / HLSMAX, lightness / HLSMAX, saturation / HLSMAX)
def tint_luminance(tint, lum):
    """Tints a HLSMAX based luminance"""
    # See: http://ciintelligence.blogspot.co.uk/2012/02/converting-excel-theme-color-and-tint.html
    if tint < 0:
        return int(round(lum * (1.0 + tint)))
    else:
        return int(round(lum * (1.0 - tint) + (HLSMAX - HLSMAX * (1.0 - tint))))
def rgb_to_ms_hls(red, green=None, blue=None) -> Tuple[int, int, int]:
    """Converts rgb values in range (0,1) or a hex string of the form '[#aa]rrggbb' to HLSMAX based HLS, (alpha values are ignored)"""
    if green is None:
        if isinstance(red, str):
            if len(red) > 6:
                red = red[-6:]  # Ignore preceding '#' and alpha values
            blue = int(red[4:], 16) / RGBMAX
            green = int(red[2:4], 16) / RGBMAX
            red = int(red[0:2], 16) / RGBMAX
        else:
            red, green, blue = red
    assert blue is not None
    h, l, s = rgb_to_hls(red, green, blue)
    return (int(round(h * HLSMAX)), int(round(l * HLSMAX)), int(round(s * HLSMAX)))
def get_theme_colors(wb):
    """Gets theme colors from the workbook"""
    # see: https://groups.google.com/forum/#!topic/openpyxl-users/I0k3TfqNLrc
    from openpyxl.xml.functions import QName, fromstring # type: ignore
    xlmns = 'http://schemas.openxmlformats.org/drawingml/2006/main'
    root = fromstring(wb.loaded_theme)
    themeEl = root.find(QName(xlmns, 'themeElements').text)
    colorSchemes = themeEl.findall(QName(xlmns, 'clrScheme').text)
    firstColorScheme = colorSchemes[0]
    colors = []
    for c in ['lt1', 'dk1', 'lt2', 'dk2', 'accent1', 'accent2', 'accent3', 'accent4', 'accent5', 'accent6']:
        accent = firstColorScheme.find(QName(xlmns, c).text)
        # if 'window' in accent.getchildren()[0].attrib['val']:
        #     colors.append(accent.getchildren()[0].attrib['lastClr'])
        # else:
        #     colors.append(accent.getchildren()[0].attrib['val'])
        colors.append(accent[0].attrib['val'])
    return colors
def theme_and_tint_to_rgb(wb, theme, tint):
    """Given a workbook, a theme number and a tint return a hex based rgb"""
    rgb = get_theme_colors(wb)[theme]
    h, l, s = rgb_to_ms_hls(rgb)
    return rgb_to_hex(*ms_hls_to_rgb(h, tint_luminance(tint, l), s))
    
    

print("0___讀取字表___")
PanCSheet = load_workbook(filename = input_path, data_only=True, keep_vba=False)

main_sheet = PanCSheet["主表-睇真尐註解"]
for n in range(main_sheet.max_column, 0, -1):
    if main_sheet.cell(row=5, column=n).value == "x":
        main_sheet.delete_cols(n)
headers      = main_sheet[2]
headers_mark = main_sheet[5]
headers_name = main_sheet[6]
assert isinstance(headers, tuple) and isinstance(headers_mark, tuple) and isinstance(headers_name, tuple)

output_dir = "Z:/Proj/Jyutdict/泛粵字表/Automatic/"
sql_ifaamjyut_header = """CREATE TABLE `IFaamjyut` (
  `id` int(11) NOT NULL,
  `col` varchar(10) CHARACTER SET utf8mb4 COLLATE utf8mb4_bin DEFAULT NULL,
  `kind` int(11) DEFAULT '0',
  `fullname` varchar(20) CHARACTER SET utf8mb4 COLLATE utf8mb4_bin DEFAULT NULL,
  `fullname_note` varchar(20) CHARACTER SET utf8mb4 COLLATE utf8mb4_bin DEFAULT NULL,
  `color` varchar(7) CHARACTER SET utf8mb4 COLLATE utf8mb4_bin DEFAULT NULL,
  `note` tinytext CHARACTER SET utf8mb4 COLLATE utf8mb4_bin
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;
ALTER TABLE `IFaamjyut` ADD PRIMARY KEY( `id`);
ALTER TABLE `IFaamjyut` ADD UNIQUE( `id`);

INSERT INTO `IFaamjyut` (`id`, `col`, `kind`, `fullname`, `fullname_note`, `color`, `note`) VALUES
"""

print("1___讀取標籤___")
for n in range(2, len(headers)):
    header_name_value = headers_name[n].value
    header_value = headers[n].value if headers[n].value else ""
    header_mark_value = int(str(headers_mark[n].value)) if headers_mark[n].value else 0
    assert header_name_value is None or isinstance(header_name_value, str)
    assert isinstance(header_value, str)
    
    if header_mark_value > 0:
        locale_name_splited: List[str] = (header_name_value+",").split(",")[0:2] if header_name_value else ["", ""]
        header_name = header_value
    else:
        locale_name_splited: List[str] = ["", ""]
        if header_name_value and "," in header_name_value:
            header_name, locale_name_splited[0] = header_name_value.split(",")
        else:
            header_name            = header_value
            locale_name_splited[0] = header_name_value if header_name_value else header_value
    HEADER_INFO_COL_NAME.append(header_name)
    HEADER_INFO_MARK.append(header_mark_value)
    HEADER_INFO_FULL_NAME.append(locale_name_splited)
    cell_fill = headers[n].fill
    if len(str(cell_fill.fgColor.rgb))==8:
        cell_color = headers[n].fill.fgColor.rgb[2:]
    else:
        cell_theme = cell_fill.start_color.theme
        cell_tint = cell_fill.start_color.tint
        cell_color = theme_and_tint_to_rgb(PanCSheet, cell_theme, cell_tint)
    HEADER_INFO_COLOR.append(cell_color)
HEADER_INFO_COL_COUNT = len(headers) # 附註列
print("2___輸出表頭___")
with open(output_dir+"IFaamjyut"+time_str+".sql", "w", encoding="utf-8") as f:
    f.write(sql_ifaamjyut_header)
    for n in range(HEADER_INFO_COL_COUNT):
        sql_row = "%s(%d, '%s', %d, '%s', '%s', '#%s', '%s')" % (
            ",\n" if n>0 else "",
            n+1, HEADER_INFO_COL_NAME[n], HEADER_INFO_MARK[n], HEADER_INFO_FULL_NAME[n][0], 
            HEADER_INFO_FULL_NAME[n][1], HEADER_INFO_COLOR[n], ""
        )
        f.write(sql_row)
    f.write(",\n(%d, '附', 0, '附註', '', '', '');" % (HEADER_INFO_COL_COUNT+1))
    
sql_row_modal = "(%d, " + ('%s, ' * (HEADER_INFO_COL_COUNT)) + "%s)"
index = 1
col_length = np.ones((HEADER_INFO_COL_COUNT+1, ), dtype="int")
main_sheet_sqls = []
dict_n2index = {}
for n in tqdm.trange(7, main_sheet.max_row+1):
    result = format_locale_record(n)
    if not result[2]:
        dict_n2index[n] = -1
        continue
    sql_row_str = sql_row_modal % tuple([index] + result[0])
    col_length = np.max(np.vstack((col_length, result[1])), axis=0)
    main_sheet_sqls.append(sql_row_str)
    dict_n2index[n] = index
    index += 1
    #if index>10: break

sql_jfaamjyut_header = "CREATE TABLE `JFaamjyut` (\n  `id` int(5) NOT NULL"
for n, i in enumerate(HEADER_INFO_COL_NAME + ["附"]):
    sql_jfaamjyut_header = sql_jfaamjyut_header + \
    ",\n  `%s` varchar(%s) CHARACTER SET utf8mb4 COLLATE utf8mb4_bin DEFAULT NULL" % \
    (i, col_length[n])
sql_jfaamjyut_header = sql_jfaamjyut_header + ",\n  PRIMARY KEY (`id`)\n) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4;\n\n"
sql_jfaamjyut_insert = "INSERT INTO `JFaamjyut` (`" + "`, `".join(["id"] + HEADER_INFO_COL_NAME + ["附"]) + "`) VALUES\n"

print("3___輸出內容___")
with open(output_dir+"JFaamjyut_1"+time_str+".sql", "w", encoding="utf-8") as f:
    f.write(sql_jfaamjyut_header)
    f.write(sql_jfaamjyut_insert)
    for n, i in enumerate(main_sheet_sqls[:3000]):
        if n>0: f.write(",\n")
        f.write(i)
    f.write(";")
if len(main_sheet_sqls)>=3000:
    with open(output_dir+"JFaamjyut_2"+time_str+".sql", "w", encoding="utf-8") as f:
        f.write(sql_jfaamjyut_insert)
        for n, i in enumerate(main_sheet_sqls[3000:]):
            if n>0: f.write(",\n")
            f.write(i)
        f.write(";")